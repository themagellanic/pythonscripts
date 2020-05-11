
  #       '''  install the built-in library using pip install openpyxl '''
		# '''	https://myaccount.google.com/lesssecureapps?pli=1 ''' 
		# '''	you have to allow less secure app acecss in order to make the script work. '''

import openpyxl
import smtplib
import getpass
import itertools

wb = openpyxl.load_workbook('sample.xlsx') # location of excel sheet.

sheet = wb.active

 
a = []
b = []
n = input("number of rows(no of students in excel sheet) ")
c1  = input("column contaning roll numbers of students ")
c2 = input("column containg marks of students ")
for i in range(1,n+1):
	x = sheet.cell(row=i,column=c1)
	z = sheet.cell(row=i,column=c2)
	
	y = x.value
	y.lower()
	y = y + '@iiitl.ac.in'
	a.append(y)
	b.append(z.value)

user = "lit2018063@iiitl.ac.in"
password = getpass.getpass()
message1 = """\
Subject:PPL midsem marks

Your midsem marks for PPL are """

message2 = """\n
Thanks & Regards
Your Name
"""
for c,d in zip(a,b):
	s = smtplib.SMTP('smtp.gmail.com',587)
	s.starttls()
	print(user,c,d ,type(user), type(c), type(d))
	s.login(user,password)
	s.sendmail(user,c,message1+ str(d) + message2)
	s.quit()


